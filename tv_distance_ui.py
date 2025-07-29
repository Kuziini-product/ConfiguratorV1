
import streamlit as st
from openpyxl import load_workbook
import io

INCH_CM = 2.54
RAPORT_16_9 = (16 / ((16**2 + 9**2)**0.5), 9 / ((16**2 + 9**2)**0.5))

def calculeaza_diagonala(d_m):
    return round(d_m * 39.37 * 0.84)

def dimensiuni_televizor(diagonala_inch):
    diagonala_cm = diagonala_inch * INCH_CM
    latime = diagonala_cm * RAPORT_16_9[0]
    inaltime = diagonala_cm * RAPORT_16_9[1]
    return round(latime / 100, 2), round(inaltime / 100, 2)

LANG = st.sidebar.selectbox("🌐 Language / Limba", ["Română", "English"])
IS_RO = LANG == "Română"

TXT = {
    "title": "📐 Configurator diagonala TV în funcție de distanță" if IS_RO else "📐 TV Diagonal Configurator by Viewing Distance",
    "distance_label": "📏 Alege distanța de vizionare (m)" if IS_RO else "📏 Choose viewing distance (m)",
    "export_button": "💾 Exportă în Excel cu aceste valori" if IS_RO else "💾 Export to Excel with these values",
    "download_label": "📥 Descarcă fișierul Excel actualizat" if IS_RO else "📥 Download updated Excel file",
    "recommend": "Kuziini recomandă" if IS_RO else "Kuziini recommends",
    "for_distance": "pentru distanța de" if IS_RO else "for a viewing distance of",
    "size": "lățime" if IS_RO else "width",
    "height": "înălțime" if IS_RO else "height",
    "tv_models": "📺 Modele TV recomandate" if IS_RO else "📺 Recommended TV Models",
    "terrace_tv": "🏡 Recomandă televizor pentru terasă" if IS_RO else "🏡 Recommend Terrace TV"
}

st.set_page_config(page_title="Kuziini TV Configurator", layout="wide")
st.image("Kuziini_logo_negru.png", width=320)
st.markdown(f"<h2 style='text-align:center; color:black;'>{TXT['title']}</h2>", unsafe_allow_html=True)

# Load workbook
wb = load_workbook("www.xlsx", data_only=True)
ws = wb.active

col1, col2 = st.columns([1, 1])

with col1:
    distanta = st.slider(TXT["distance_label"], 1.0, 5.0, 2.5, 0.1)
    diagonala_inch = calculeaza_diagonala(distanta)
    latime_m, inaltime_m = dimensiuni_televizor(diagonala_inch)

    show_terrace = st.checkbox(TXT["terrace_tv"])

    html_box = f"""
<div style="background-color:#F0F9FF;padding:1.5rem;border-radius:12px;
            border:2px solid #0B5394;text-align:center;">
    <h1 style="color:#FF5722;font-size:3rem;">{diagonala_inch}"</h1>
    <h3 style="color:#0B5394;">{TXT['recommend']}</h3>
    <p>{TXT['for_distance']} {distanta} m</p>
    <p style="font-weight:bold;">🖼️ {TXT['size']} {latime_m} m × {TXT['height']} {inaltime_m} m</p>
</div>
"""
    st.markdown(html_box, unsafe_allow_html=True)

    if st.button(TXT["export_button"]):
        ws["B1"] = distanta
        ws["B6"] = round(distanta / 30, 2)
        ws["B7"] = round(distanta / 25, 2)
        output = io.BytesIO()
        wb.save(output)
        st.download_button(TXT["download_label"], output.getvalue(), "recomandare_tv_actualizat.xlsx")

with col2:
    st.image("TV.png", caption="Kuziini × Samsung", use_container_width=True)
    st.markdown("<p style='text-align:center;font-weight:bold;color:black;'>Living Kuziini × Samsung</p>", unsafe_allow_html=True)


# 📺 MODELE TV RECOMANDATE (AFIȘATE ÎNTRE EXPORT ȘI IMAGINE)

st.markdown(f"<h4 style='margin-top:2rem;'>{TXT['tv_models']}</h4>", unsafe_allow_html=True)

with st.container():
    st.markdown("""
    <style>
        .tv-compare {
            display: flex;
            flex-direction: row;
            flex-wrap: wrap;
            justify-content: space-between;
            gap: 1rem;
        }
        .tv-box {
            flex: 1 1 45%;
            background-color: #F9F9F9;
            padding: 1rem;
            border: 1px solid #CCC;
            border-radius: 12px;
            min-width: 240px;
        }
        @media screen and (max-width: 768px) {
            .tv-compare {
                flex-direction: column;
            }
        }
    </style>
    """, unsafe_allow_html=True)

    if show_terrace:
        st.markdown("""
        <div class='tv-compare'>
            <div class='tv-box'>
                <h4>Samsung 55LST7T – The Terrace</h4>
                <ul>
                    <li>Outdoor TV</li>
                    <li>UHD 4K</li>
                    <li>Ultra Bright Picture Quality</li>
                </ul>
                <a href='https://www.samsung.com/ro/lifestyle-tvs/all-lifestyle-tvs/?the-terrace' target='_blank'>Vezi mai multe</a>
            </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        if diagonala_inch <= 55:
            models = [
                ("Samsung AU7092", ["4K Crystal UHD", "Smart Hub", "HDR10+"], "https://www.samsung.com/ro/tvs/all-tvs/?qled-tv+in-stock"),
                ("Samsung Q60B", ["QLED", "AirSlim", "Dual LED"], "https://www.samsung.com/ro/tvs/all-tvs/?qled-tv+in-stock"),
            ]
        elif 55 < diagonala_inch <= 75:
            models = [
                ("Samsung QN85C", ["Neo QLED 4K", "Mini LED", "Quantum HDR"], "https://www.samsung.com/ro/tvs/all-tvs/?neo-qled-tv+uhd-4k+in-stock"),
                ("Samsung QN90B", ["144Hz Gaming", "Dolby Atmos", "Ultra Viewing Angle"], "https://www.samsung.com/ro/tvs/all-tvs/?neo-qled-tv+uhd-4k+in-stock"),
            ]
        else:
            models = [
                ("Samsung QN95C", ["Neo QLED 4K", "Precision Contrast", "Slim One Connect"], "https://www.samsung.com/ro/tvs/all-tvs/?neo-qled-tv+uhd-8k+in-stock"),
                ("Samsung S95C OLED", ["OLED 4K", "Quantum HDR OLED+", "Dolby Atmos"], "https://www.samsung.com/ro/tvs/all-tvs/?oled-tv+in-stock"),
            ]

        html_blocks = []
        for name, features, link in models:
            feats = "".join([f"<li>✔️ {f}</li>" for f in features])
            html_blocks.append(f"""
            <div class='tv-box'>
                <h4><a href='{link}' target='_blank'>{name}</a></h4>
                <ul>{feats}</ul>
            </div>
            """)

        st.markdown(f"""<div class='tv-compare'>{''.join(html_blocks)}</div>""", unsafe_allow_html=True)

# Imagine Kuziini TV

st.markdown(f"<h4 style='margin-top:3rem;'>{TXT['tv_models']}</h4>", unsafe_allow_html=True)

if show_terrace:
    st.markdown("### Samsung 55LST7T – The Terrace")
    st.markdown("• Outdoor TV • UHD 4K • Ultra Bright Picture Quality")
    st.markdown("[Vezi mai multe](https://www.samsung.com/ro/lifestyle-tvs/all-lifestyle-tvs/?the-terrace)", unsafe_allow_html=True)
else:
    if diagonala_inch <= 55:
        models = [
            ("Samsung AU7092", ["4K Crystal UHD", "Smart Hub", "HDR10+"], "https://www.samsung.com/ro/tvs/all-tvs/?qled-tv+in-stock"),
            ("Samsung Q60B", ["QLED", "AirSlim", "Dual LED"], "https://www.samsung.com/ro/tvs/all-tvs/?qled-tv+in-stock"),
        ]
    elif 55 < diagonala_inch <= 75:
        models = [
            ("Samsung QN85C", ["Neo QLED 4K", "Mini LED", "Quantum HDR"], "https://www.samsung.com/ro/tvs/all-tvs/?neo-qled-tv+uhd-4k+in-stock"),
            ("Samsung QN90B", ["144Hz Gaming", "Dolby Atmos", "Ultra Viewing Angle"], "https://www.samsung.com/ro/tvs/all-tvs/?neo-qled-tv+uhd-4k+in-stock"),
        ]
    else:
        models = [
            ("Samsung QN95C", ["Neo QLED 4K", "Precision Contrast", "Slim One Connect"], "https://www.samsung.com/ro/tvs/all-tvs/?neo-qled-tv+uhd-8k+in-stock"),
            ("Samsung S95C OLED", ["OLED 4K", "Quantum HDR OLED+", "Dolby Atmos"], "https://www.samsung.com/ro/tvs/all-tvs/?oled-tv+in-stock"),
        ]

    col1, col2 = st.columns(2)
    for col, (name, features, link) in zip([col1, col2], models):
        with col:
            st.markdown(f"#### [{name}]({link})", unsafe_allow_html=True)
            for f in features:
                st.markdown(f"✅ {f}")
